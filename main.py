import telebot
from telebot import types
import random
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import PatternFill
import emoji
import os.path
import time 


def saving(item):
    count = 2
    workbook = load_workbook(filename=file_name)
    sheet = workbook['Sheet']
    sand_fill = PatternFill(start_color='f7e899',
                   end_color='f7e899',
                   fill_type='solid')
    last_row = 1
    for row in sheet.iter_rows():
        if row[0].value is None:
            break
        last_row += 1
    sheet.cell(row=last_row,column = 1).value = last_row-1    
    while count != 8:
        if item.find(',') != -1:
            this = item[:item.find(',')]
            print(this)
            sheet.cell(row=last_row,column = count).value = this
            
            count += 1
            item = item[item.find(',')+1:]
        else : 
            this = item    
            print(this)
            sheet.cell(row=last_row,column = count).value = this
            
            count += 1
    for row in sheet.rows:
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = sand_fill
    
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
    workbook.save(f'{file_name}')
        
bot = telebot.TeleBot("6000528616:AAG-vFtWb6-DDQKyOwyRf9TlN6ejOYub_dw")

@bot.message_handler(commands=['start'])
def start(message):
    global markup
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton(f"Внести дані {emoji.emojize(':bookmark_tabs:')}")   
    btn2 = types.KeyboardButton(f"Отримати звіт {emoji.emojize(':black_square_button:')}")
    btn3 = types.KeyboardButton(f"Обнулити звіт {emoji.emojize(':cross_mark:')}")
    markup.add(btn1,btn2,btn3)
    user_id = message.from_user.id
    ans = f'Обери дію'
    global file_name
    file_name = f'./report_{user_id}.xlsx'
    check_file = os.path.exists(file_name)
    if check_file:
        bot.send_message(message.chat.id,f'Вітаю, <b>{message.from_user.first_name}</b>!\nExcel файл знайдено!',parse_mode='html')
        time.sleep(2)
        bot.send_message(message.chat.id, ans, reply_markup=markup, parse_mode='html')
    else: 
        bot.send_message(message.chat.id,f'Вітаю, <b>{message.from_user.first_name}</b>!\nExcel файл не знайдено',parse_mode='html')
        time.sleep(1)
        bot.send_message(message.chat.id,'Генерую Excel файл')
        report = openpyxl.Workbook()
        sheet = report.active
        
        sheet.cell(row=1,column = 1).value = "№"
        sheet.cell(row=1,column = 2).value = "Прізвище ім'я клієнта"
        sheet.cell(row=1,column = 3).value = 'Дата продажу'
        sheet.cell(row=1,column = 4).value = 'Номер телефону клієнта'
        sheet.cell(row=1,column = 5).value = 'Спосіб допродажу'
        sheet.cell(row=1,column = 6).value = 'Джерело допродажу'
        sheet.cell(row=1,column = 7).value = 'Сума допродажу'
        
        report.save(f'{file_name}')
        time.sleep(3)
        bot.send_message(message.chat.id,'Файл згенеровано! Продовжуємо :)', reply_markup=markup, parse_mode='html')

@bot.message_handler(content_types=['text'])
def imp(message):
    markep = types.ReplyKeyboardMarkup(resize_keyboard=True)   
    btn1 = types.KeyboardButton(f"Yes {emoji.emojize(':check_mark:')}")
    btn2 = types.KeyboardButton(f"No {emoji.emojize(':cross_mark:')}")
    markep.add(btn1,btn2)
    if(message.text == f"Внести дані {emoji.emojize(':bookmark_tabs:')}"):
        bot.send_message(message.chat.id, text = f"<b>Введи через кому і без пробілів!</b>\nПрізвище та ім'я клієнта\nДату допродажу\nНомер телефону клієнта\nСпосіб допродажу\nДжерело допродажу\nСума допродажу",parse_mode='html')
        bot.send_message(message.chat.id, text = "Приклад\nСічка Тетяна,09/01/22,380954487728,за промокодом,116066,938")
    elif(message.text.count(',') == 5):
        time.sleep(1)
        bot.send_message(message.chat.id,text='Прийнято :)', reply_markup=markup, parse_mode='html')
        item = message.text
        saving(item)
    elif(message.text == f"Отримати звіт {emoji.emojize(':black_square_button:')}"):
        bot.send_message(message.chat.id, text = 'Твій згенерований звіт:')
        bot.send_document(message.chat.id, open(file_name, 'rb'))
    elif(message.text == f"Обнулити звіт {emoji.emojize(':cross_mark:')}"):
        bot.send_message(message.chat.id, text = 'Ти впевнений, що хочеш обнулити таблицю?', reply_markup =markep)
    elif(message.text == f"Yes {emoji.emojize(':check_mark:')}"):
        report = openpyxl.Workbook()
        sheet = report.active
        
        sheet.cell(row=1,column = 1).value = "№"
        sheet.cell(row=1,column = 2).value = "Прізвище ім'я клієнта"
        sheet.cell(row=1,column = 3).value = 'Дата продажу'
        sheet.cell(row=1,column = 4).value = 'Номер телефону клієнта'
        sheet.cell(row=1,column = 5).value = 'Спосіб допродажу'
        sheet.cell(row=1,column = 6).value = 'Джерело допродажу'
        sheet.cell(row=1,column = 7).value = 'Сума допродажу'
        
        report.save(f'{file_name}')
        time.sleep(3)
        bot.send_message(message.chat.id,'Звіт обнулився! Продовжуємо :)', reply_markup=markup, parse_mode='html')
    
    elif(message.text == f"No {emoji.emojize(':cross_mark:')}"):
        bot.send_message(message.chat.id,text='Прийнято :)', reply_markup=markup, parse_mode='html')

    else: 
        bot.send_sticker(message.chat.id, 'CAACAgIAAxkBAAEH_D9kAmwoIL5KsxWxEGUHtROI9XJuMgACFwEAAjDUnRHjh3RDGOQJYS4E')
        bot.send_message(message.chat.id, 'Вибачай, я ще такого не розумію :)')


bot.polling(none_stop=True)